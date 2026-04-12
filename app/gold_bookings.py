"""Gold booking records: pending until 25% advance (configurable) is marked paid.

Stored in SQLite (same DB file as customers). Optional sync to .xlsx on disk.
"""

from __future__ import annotations

import json
import logging
import re
import sqlite3
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.config import GOLD_BOOKING_ADVANCE_PCT, GOLD_BOOKINGS_XLSX_PATH
from app.gold_rates import get_rates
from app.jewellery_quote import QuoteParams, compute_breakdown

logger = logging.getLogger(__name__)

DB_PATH = Path(__file__).resolve().parent.parent / "customers.db"
_conn = None

STATUS_PENDING = "pending_advance"
STATUS_CONFIRMED = "confirmed"
STATUS_CANCELLED = "cancelled"

_GOLD_BOOKING_BLOCK = re.compile(
    r"\[GOLD_BOOKING\](.*?)\[/GOLD_BOOKING\]",
    re.IGNORECASE | re.DOTALL,
)

_BOOK_CMD = re.compile(
    r"^\s*book\s+(?:gold\s+)?(\d+(?:\.\d+)?)\s*g?\s+(\d{1,2})\s*k?\s*(.*)$",
    re.IGNORECASE,
)

def _get_conn():
    global _conn
    if _conn is None:
        _conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        _conn.row_factory = sqlite3.Row
        _conn.execute("PRAGMA journal_mode=WAL")
        _conn.execute("""
            CREATE TABLE IF NOT EXISTS gold_bookings (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                phone           TEXT NOT NULL,
                customer_name   TEXT NOT NULL DEFAULT '',
                grams           REAL NOT NULL,
                karat           INTEGER NOT NULL,
                prompt          TEXT NOT NULL DEFAULT '',
                gold_value_inr  REAL NOT NULL,
                making_inr      REAL NOT NULL,
                subtotal_inr    REAL NOT NULL,
                sale_gst_inr    REAL NOT NULL,
                grand_total_inr REAL NOT NULL,
                advance_pct     REAL NOT NULL,
                advance_inr     REAL NOT NULL,
                advance_paid    INTEGER NOT NULL DEFAULT 0,
                status          TEXT NOT NULL DEFAULT 'pending_advance',
                created_at      TEXT NOT NULL,
                confirmed_at    TEXT,
                rate_snapshot   TEXT NOT NULL DEFAULT '{}'
            )
        """)
        _conn.commit()
        logger.info("gold_bookings table ready")
    return _conn


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def strip_gold_booking_tag(reply_text: str) -> tuple[str, str | None]:
    m = _GOLD_BOOKING_BLOCK.search(reply_text or "")
    if not m:
        return reply_text or "", None
    inner = m.group(1).strip()
    cleaned = (reply_text[: m.start()] + reply_text[m.end() :]).strip()
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned, inner


def parse_booking_tag_inner(inner: str) -> tuple[float, int, str] | None:
    """Returns (grams, karat, prompt) or None."""
    raw = (inner or "").strip()
    if not raw:
        return None
    grams = None
    karat = 22
    prompt = ""
    # quoted prompt
    mq = re.search(r'prompt\s*=\s*"([^"]*)"', raw, re.I)
    if mq:
        prompt = mq.group(1).strip()
        raw = raw[: mq.start()] + raw[mq.end() :]
    for part in re.split(r"[\s,;]+", raw):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        kl = k.strip().lower()
        v = v.strip().strip('"').strip("'")
        if kl in ("grams", "gram", "g", "weight", "w"):
            try:
                grams = float(re.sub(r"[^\d.]", "", v))
            except ValueError:
                pass
        elif kl in ("karat", "k", "purity"):
            try:
                karat = int(re.sub(r"[^\d]", "", v))
            except ValueError:
                pass
        elif kl == "prompt" and not prompt:
            prompt = v
    if grams is None or grams <= 0 or grams > 10_000:
        return None
    if karat not in (24, 22, 18, 14):
        karat = 22
    return grams, karat, prompt[:2000]


def parse_book_command(text: str) -> tuple[float, int, str] | None:
    m = _BOOK_CMD.match((text or "").strip())
    if not m:
        return None
    w = float(m.group(1))
    k = int(m.group(2))
    prompt = (m.group(3) or "").strip()
    if w <= 0 or w > 10_000 or k not in (24, 22, 18, 14):
        return None
    return w, k, prompt[:2000]


def parse_owner_booking_confirm(text: str) -> int | None:
    t = (text or "").strip().lower()
    patterns = (
        r"^advance\s+paid\s+(\d+)\s*$",
        r"^advance\s+(\d+)\s*$",
        r"^booking\s+confirm\s+(\d+)\s*$",
        r"^confirm\s+booking\s+(\d+)\s*$",
    )
    for p in patterns:
        m = re.match(p, t)
        if m:
            return int(m.group(1))
    return None


async def create_booking(
    phone: str,
    customer_name: str,
    grams: float,
    karat: int,
    prompt: str,
) -> dict | None:
    rates = await get_rates()
    b = compute_breakdown(rates, QuoteParams(weight_grams=grams, purity_karat=karat, discount_percent=0.0))
    if not b:
        return None
    advance_pct = GOLD_BOOKING_ADVANCE_PCT
    advance_inr = round(b["subtotal"] * (advance_pct / 100.0), 2)
    snap = json.dumps(
        {
            "gold_24k_per_gram": rates.get("gold_24k_per_gram"),
            "gold_22k_per_gram": rates.get("gold_22k_per_gram"),
            "import_duty_pct": rates.get("import_duty_pct"),
            "gst_note_pct": rates.get("gst_pct"),
            "source": rates.get("source"),
        },
        ensure_ascii=False,
    )
    conn = _get_conn()
    cur = conn.execute(
        """
        INSERT INTO gold_bookings (
            phone, customer_name, grams, karat, prompt,
            gold_value_inr, making_inr, subtotal_inr, sale_gst_inr, grand_total_inr,
            advance_pct, advance_inr, advance_paid, status, created_at, rate_snapshot
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?)
        """,
        (
            phone,
            customer_name or "",
            grams,
            karat,
            prompt or "",
            b["gold_value"],
            b["making_amount"],
            b["subtotal"],
            b["sale_gst_amount"],
            b["grand_total"],
            advance_pct,
            advance_inr,
            STATUS_PENDING,
            _now_iso(),
            snap,
        ),
    )
    conn.commit()
    bid = int(cur.lastrowid)
    row = get_booking(bid)
    _maybe_sync_xlsx()
    logger.info("Gold booking #%s created phone=%s g=%s k=%s", bid, phone, grams, karat)
    return row


def get_booking(booking_id: int) -> dict | None:
    conn = _get_conn()
    r = conn.execute("SELECT * FROM gold_bookings WHERE id = ?", (booking_id,)).fetchone()
    return dict(r) if r else None


def count_bookings_by_status(status: str) -> int:
    conn = _get_conn()
    row = conn.execute(
        "SELECT COUNT(*) AS c FROM gold_bookings WHERE status = ?",
        (status,),
    ).fetchone()
    return int(row["c"]) if row else 0


def list_bookings(limit: int = 500, status: str | None = None, order_desc: bool = True) -> list[dict]:
    conn = _get_conn()
    order = "DESC" if order_desc else "ASC"
    if status:
        rows = conn.execute(
            f"SELECT * FROM gold_bookings WHERE status = ? ORDER BY id {order} LIMIT ?",
            (status, limit),
        ).fetchall()
    else:
        rows = conn.execute(
            f"SELECT * FROM gold_bookings ORDER BY id {order} LIMIT ?",
            (limit,),
        ).fetchall()
    return [dict(x) for x in rows]


def mark_advance_paid(booking_id: int) -> dict | None:
    row = get_booking(booking_id)
    if not row:
        return None
    if row["status"] != STATUS_PENDING:
        return row
    conn = _get_conn()
    cur = conn.execute(
        """
        UPDATE gold_bookings
        SET advance_paid = 1, status = ?, confirmed_at = ?
        WHERE id = ? AND status = ?
        """,
        (STATUS_CONFIRMED, _now_iso(), booking_id, STATUS_PENDING),
    )
    conn.commit()
    if cur.rowcount == 0:
        return get_booking(booking_id)
    _maybe_sync_xlsx()
    logger.info("Gold booking #%s advance marked paid", booking_id)
    return get_booking(booking_id)


def bookings_to_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gold Bookings"
    headers = [
        "id",
        "created_at",
        "phone",
        "customer_name",
        "grams",
        "karat",
        "prompt",
        "gold_value_inr",
        "making_inr",
        "subtotal_inr",
        "sale_gst_inr",
        "grand_total_inr",
        "advance_pct",
        "advance_inr",
        "advance_paid",
        "status",
        "confirmed_at",
    ]
    ws.append(headers)
    for b in list_bookings(limit=10_000, order_desc=False):
        ws.append(
            [
                b["id"],
                b["created_at"],
                b["phone"],
                b["customer_name"],
                b["grams"],
                b["karat"],
                b["prompt"],
                b["gold_value_inr"],
                b["making_inr"],
                b["subtotal_inr"],
                b["sale_gst_inr"],
                b["grand_total_inr"],
                b["advance_pct"],
                b["advance_inr"],
                "yes" if b["advance_paid"] else "no",
                b["status"],
                b["confirmed_at"] or "",
            ]
        )
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["G"].width = 40
    return wb


def export_bookings_xlsx_bytes() -> bytes:
    wb = bookings_to_workbook()
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _maybe_sync_xlsx() -> None:
    path = (GOLD_BOOKINGS_XLSX_PATH or "").strip()
    if not path:
        return
    try:
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb = bookings_to_workbook()
        wb.save(str(p))
        logger.info("Synced gold bookings workbook to %s", p)
    except Exception:
        logger.exception("Failed to sync gold bookings xlsx to %s", path)


def format_booking_created_customer(lang: str, b: dict) -> str:
    if lang == "en":
        return (
            f"📋 *Gold booking registered — #{b['id']}*\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 {b['customer_name'] or 'Customer'}\n"
            f"⚖️ *{b['grams']:.2f} g*  {b['karat']}K gold\n"
            f"📝 {b['prompt'] or '—'}\n\n"
            f"🪙 Gold (today’s rate): ₹{b['gold_value_inr']:,.0f}\n"
            f"🔧 Making: ₹{b['making_inr']:,.0f}\n"
            f"📊 *Booking value (gold + making): ₹{b['subtotal_inr']:,.0f}*\n"
            f"🧾 + Sale GST (est.): ₹{b['sale_gst_inr']:,.0f}\n"
            f"💰 Indicative total: ₹{b['grand_total_inr']:,.0f}\n\n"
            f"✅ *Advance to confirm:* *{b['advance_pct']:.0f}%* = *₹{b['advance_inr']:,.0f}*\n"
            "Booking is *pending* until this advance is received at the store / as instructed.\n\n"
            "📞 Sharda Jewellers, Bemetara: +91 94255 61850, +91 70003 44110"
        )
    if lang == "hinglish":
        return (
            f"📋 *Gold booking register — #{b['id']}*\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 {b['customer_name'] or 'Customer'}\n"
            f"⚖️ *{b['grams']:.2f} g*  {b['karat']}K\n"
            f"📝 {b['prompt'] or '—'}\n\n"
            f"🪙 Gold (aaj ka rate): ₹{b['gold_value_inr']:,.0f}\n"
            f"🔧 Making: ₹{b['making_inr']:,.0f}\n"
            f"📊 *Booking value (gold+making): ₹{b['subtotal_inr']:,.0f}*\n"
            f"🧾 Sale GST (approx): ₹{b['sale_gst_inr']:,.0f}\n"
            f"💰 Indicative total: ₹{b['grand_total_inr']:,.0f}\n\n"
            f"✅ *Confirm karne ke liye advance:* *{b['advance_pct']:.0f}%* = *₹{b['advance_inr']:,.0f}*\n"
            "Advance milne tak booking *pending* rahegi.\n\n"
            "📞 Sharda Jewellers Bemetara: +91 94255 61850, +91 70003 44110"
        )
    return (
        f"📋 *सोना बुकिंग दर्ज — #{b['id']}*\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 {b['customer_name'] or 'ग्राहक'}\n"
        f"⚖️ *{b['grams']:.2f} ग्राम*  {b['karat']}K\n"
        f"📝 {b['prompt'] or '—'}\n\n"
        f"🪙 सोना (आज का भाव): ₹{b['gold_value_inr']:,.0f}\n"
        f"🔧 मेकिंग: ₹{b['making_inr']:,.0f}\n"
        f"📊 *बुकिंग मूल्य (सोना+मेकिंग): ₹{b['subtotal_inr']:,.0f}*\n"
        f"🧾 बिक्री GST (अनुमान): ₹{b['sale_gst_inr']:,.0f}\n"
        f"💰 संकेतक कुल: ₹{b['grand_total_inr']:,.0f}\n\n"
        f"✅ *पुष्टि हेतु एडवांस:* *{b['advance_pct']:.0f}%* = *₹{b['advance_inr']:,.0f}*\n"
        "एडवांस मिलने तक बुकिंग *लंबित* रहेगी।\n\n"
        "📞 शारदा ज्वेलर्स, बेमेतरा: +91 94255 61850, +91 70003 44110"
    )


def format_booking_confirmed_customer(lang: str, b: dict) -> str:
    if lang == "en":
        return (
            f"🎉 *Booking #{b['id']} confirmed!*\n"
            f"We’ve recorded your *{b['advance_pct']:.0f}%* advance (₹{b['advance_inr']:,.0f}).\n"
            "Thank you — Sharda Jewellers will coordinate next steps on WhatsApp or at the store."
        )
    if lang == "hinglish":
        return (
            f"🎉 *Booking #{b['id']} confirm ho gayi!*\n"
            f"*₹{b['advance_inr']:,.0f}* ({b['advance_pct']:.0f}% advance) note kar liya.\n"
            "Dhanyavaad — Sharda Jewellers aage steps batayega."
        )
    return (
        f"🎉 *बुकिंग #{b['id']} पुष्ट!*\n"
        f"आपका *{b['advance_pct']:.0f}%* एडवांस (₹{b['advance_inr']:,.0f}) दर्ज कर लिया गया।\n"
        "धन्यवाद — शारदा ज्वेलर्स आगे की जानकारी देगा।"
    )


def format_owner_new_booking(b: dict) -> str:
    return (
        f"🆕 *नई गोल्ड बुकिंग #{b['id']}*\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 {b['customer_name'] or '—'}  |  📱 wa.me/{b['phone']}\n"
        f"⚖️ {b['grams']:.2f} g  {b['karat']}K\n"
        f"📝 {b['prompt'] or '—'}\n"
        f"📊 Subtotal: ₹{b['subtotal_inr']:,.0f}  |  Advance ({b['advance_pct']:.0f}%): *₹{b['advance_inr']:,.0f}*\n"
        f"📌 Status: *pending advance*\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"Advance मिलने पर लिखें: `advance {b['id']}` या `booking confirm {b['id']}`"
    )
